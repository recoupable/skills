#!/usr/bin/env python3
"""Normalize first-pass provider CSV/TSV exports into royalty-ledger.csv.

Returns status="ok" only when financial fields (gross_amount, owner_net_amount,
period_start) are populated above 50% across rows. Lower coverage returns
status="partial" with a warnings array enumerating which expected source columns
were not found, so callers do not silently proceed with a useless ledger.
"""

from __future__ import annotations

import argparse
import calendar
import csv
import io
import json
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Callable


CANONICAL_COLUMNS = [
    "ledger_line_id",
    "catalog_asset_id",
    "source_file",
    "provider",
    "period_start",
    "period_end",
    "payment_date",
    "rights_type",
    "income_type",
    "territory",
    "platform_or_licensee",
    "gross_amount",
    "deductions",
    "participant_share",
    "owner_net_amount",
    "currency",
    "fx_rate",
    "pro_use_type",
    "pro_credits",
    "pro_bonus_type",
    "cue_sheet_ref",
    "match_confidence",
    "notes",
]


CRITICAL_FIELDS = ("gross_amount", "owner_net_amount", "period_start")
PARTIAL_THRESHOLD = 0.5
NUMERIC_FIELDS = {"gross_amount", "deductions", "owner_net_amount"}


# Per-provider source-column aliases for the financially critical canonical
# fields. Used to compute "expected columns not found" warnings when coverage
# is partial. Compared case-insensitively against input headers.
FINANCIAL_ALIASES: dict[str, dict[str, list[str]]] = {
    "ascap": {
        "gross_amount": ["Gross Amount", "Gross", "Royalty Amount"],
        "owner_net_amount": ["Publisher Share", "Net", "Royalty Amount"],
        "period_start": ["Performance Start", "Period Start"],
        "period_end": ["Performance End", "Period End"],
        "platform_or_licensee": ["Licensee", "Station", "Network", "Series / Film Title"],
    },
    "bmi": {
        "gross_amount": ["Amount", "Gross Amount", "Amount Paid"],
        "owner_net_amount": ["Publisher Net", "Net", "Amount Paid"],
        "period_start": ["Period Start", "Statement Qtr"],
        "period_end": ["Period End", "Statement Qtr"],
        "platform_or_licensee": ["Station", "Licensee", "Network"],
    },
    "mlc": {
        "gross_amount": ["Mechanical Gross", "Gross"],
        "owner_net_amount": ["Publisher Net", "Net", "Net Payable"],
        "period_start": ["Period Start", "Usage Period", "Period"],
        "period_end": ["Period End", "Usage Period", "Period"],
        "platform_or_licensee": ["DSP", "Platform", "Source"],
    },
    "distributor": {
        "gross_amount": ["Gross", "Gross Amount"],
        "owner_net_amount": ["Net", "Net Amount", "Net Payable"],
        "period_start": ["Period Start", "Sales Month", "Period"],
        "period_end": ["Period End", "Sales Month", "Period"],
        "platform_or_licensee": ["Store", "DSP", "Platform", "Source"],
    },
    "publisher-admin": {
        "gross_amount": ["Gross", "Gross Amount"],
        "owner_net_amount": ["Net", "Net Amount", "Net Payable"],
        "period_start": ["Period Start", "Sales Month", "Period"],
        "period_end": ["Period End", "Sales Month", "Period"],
        "platform_or_licensee": ["Store", "DSP", "Platform", "Source"],
    },
    "soundexchange": {
        "gross_amount": ["Gross", "Gross Amount"],
        "owner_net_amount": ["Net", "Net Amount", "Net Payable"],
        "period_start": ["Period Start", "Sales Month", "Period"],
        "period_end": ["Period End", "Sales Month", "Period"],
        "platform_or_licensee": ["Store", "DSP", "Platform", "Source"],
    },
    "direct-sync": {
        "gross_amount": ["Gross", "Gross Amount"],
        "owner_net_amount": ["Net", "Net Amount", "Net Payable"],
        "period_start": ["Period Start", "Sales Month", "Period"],
        "period_end": ["Period End", "Sales Month", "Period"],
        "platform_or_licensee": ["Store", "DSP", "Platform", "Source"],
    },
    "youtube-content-id": {
        "gross_amount": ["Gross Revenue", "Gross", "youtube_revenue_split"],
        "owner_net_amount": ["Net Revenue", "Net", "partner_revenue_usd"],
        "period_start": ["Reporting Month", "Month", "date"],
        "period_end": ["Reporting Month", "Month", "date"],
        "platform_or_licensee": [],
    },
    "curve": {
        "gross_amount": ["Gross"],
        "owner_net_amount": ["Net Payable", "Net"],
        "period_start": ["Period Start"],
        "period_end": ["Period End"],
        "platform_or_licensee": ["Royalty Source", "Source"],
    },
}


def clean_amount(value: str | None) -> str:
    if not value:
        return ""
    cleaned = value.strip().replace("$", "").replace(",", "")
    is_negative = cleaned.startswith("(") and cleaned.endswith(")")
    cleaned = cleaned.strip("()")
    if not cleaned:
        return ""
    try:
        amount = float(cleaned)
    except ValueError:
        return ""
    if is_negative:
        amount *= -1
    return f"{amount:.2f}"


def parse_month(value: str | None) -> tuple[str, str]:
    if not value:
        return "", ""
    text = value.strip()
    match = re.fullmatch(r"(\d{4})-(\d{2})", text)
    if not match:
        return text, text
    year = int(match.group(1))
    month = int(match.group(2))
    last_day = calendar.monthrange(year, month)[1]
    return f"{year:04d}-{month:02d}-01", f"{year:04d}-{month:02d}-{last_day:02d}"


def parse_quarter(value: str | None) -> tuple[str, str]:
    if not value:
        return "", ""
    text = value.strip()
    match = re.fullmatch(r"(\d{4})Q([1-4])", text)
    if not match:
        return "", ""
    year = int(match.group(1))
    quarter = int(match.group(2))
    start_month = (quarter - 1) * 3 + 1
    end_month = start_month + 2
    last_day = calendar.monthrange(year, end_month)[1]
    return f"{year:04d}-{start_month:02d}-01", f"{year:04d}-{end_month:02d}-{last_day:02d}"


def parse_yyyymmdd(value: str | None) -> str:
    if not value:
        return ""
    text = value.strip()
    match = re.fullmatch(r"(\d{4})(\d{2})(\d{2})", text)
    if not match:
        return text
    return f"{match.group(1)}-{match.group(2)}-{match.group(3)}"


def first(row: dict[str, str], *names: str) -> str:
    lower_lookup = {key.lower(): value for key, value in row.items() if isinstance(key, str)}
    for name in names:
        value = row.get(name)
        if value not in (None, ""):
            return value.strip()
        value = lower_lookup.get(name.lower())
        if value not in (None, ""):
            return value.strip()
    return ""


def has_meaningful_value(row: dict[str, str]) -> bool:
    values = [value.strip() for key, value in row.items() if isinstance(key, str) and value]
    if not values:
        return False
    first_value = values[0].lower()
    if first_value.startswith("totals") or first_value.startswith("note:"):
        return False
    return True


def asset_id(prefix: str, value: str, fallback: str) -> str:
    if value:
        return f"{prefix}:{value}"
    return f"title:{fallback.strip().lower().replace(' ', '-')}" if fallback else "unknown"


def blank_row(index: int, source_file: str) -> dict[str, str]:
    row = {column: "" for column in CANONICAL_COLUMNS}
    row["ledger_line_id"] = f"RL-{index:06d}"
    row["source_file"] = source_file
    row["match_confidence"] = "medium"
    return row


@dataclass(frozen=True)
class ProviderProfile:
    provider_name: str
    rights_type: str
    income_type: str
    normalize: Callable[[dict[str, str], int, str], dict[str, str]]


def normalize_ascap(row: dict[str, str], index: int, source_file: str) -> dict[str, str]:
    output = blank_row(index, source_file)
    title = first(row, "Work Title", "Title")
    gross = clean_amount(first(row, "Gross Amount", "Gross", "Royalty Amount"))
    net = clean_amount(first(row, "Publisher Share", "Net", "Royalty Amount"))
    output.update(
        {
            "provider": "ASCAP",
            "rights_type": "publishing",
            "income_type": "performance",
            "catalog_asset_id": asset_id("iswc", first(row, "ISWC"), title),
            "period_start": first(row, "Performance Start", "Period Start"),
            "period_end": first(row, "Performance End", "Period End"),
            "payment_date": first(row, "Distribution Date", "Payment Date"),
            "territory": first(row, "Territory"),
            "platform_or_licensee": first(row, "Licensee", "Station", "Network", "Series / Film Title"),
            "gross_amount": gross,
            "deductions": clean_amount(first(row, "Deductions", "Fees")),
            "participant_share": first(row, "Share %"),
            "owner_net_amount": net,
            "currency": first(row, "Currency") or "USD",
            "pro_use_type": first(row, "Use Type", "Perf. Type"),
            "pro_credits": first(row, "Credits", "Total Credits", "Premium Credits"),
            "pro_bonus_type": first(row, "Bonus Type", "Premium Type"),
            "cue_sheet_ref": first(row, "Cue Sheet", "Program"),
            "notes": f"work_id={first(row, 'Work ID')}" if first(row, "Work ID") else "",
        }
    )
    return output


def normalize_bmi(row: dict[str, str], index: int, source_file: str) -> dict[str, str]:
    output = blank_row(index, source_file)
    title = first(row, "Title", "Work Title")
    quarter_start, quarter_end = parse_quarter(first(row, "Statement Qtr"))
    hit_song_bonus = clean_amount(first(row, "Hit Song Bonus"))
    bonus_type = first(row, "Bonus Category", "Bonus Type") or ("Hit Song Bonus" if hit_song_bonus not in ("", "0.00") else "")
    net = clean_amount(first(row, "Publisher Net", "Net", "Amount Paid"))
    output.update(
        {
            "provider": "BMI",
            "rights_type": "publishing",
            "income_type": "performance",
            "catalog_asset_id": asset_id("iswc", first(row, "ISWC"), title),
            "period_start": first(row, "Period Start") or quarter_start,
            "period_end": first(row, "Period End") or quarter_end,
            "payment_date": first(row, "Payment Date"),
            "territory": first(row, "Territory"),
            "platform_or_licensee": first(row, "Station", "Licensee", "Network"),
            "gross_amount": clean_amount(first(row, "Amount", "Gross Amount", "Amount Paid")),
            "deductions": clean_amount(first(row, "Admin Fee", "Deductions")),
            "owner_net_amount": net,
            "currency": first(row, "Currency") or "USD",
            "pro_use_type": first(row, "Performance Type", "Use Type"),
            "pro_bonus_type": bonus_type,
            "notes": f"bmi_work_id={first(row, 'BMI Work #')}" if first(row, "BMI Work #") else "",
        }
    )
    return output


def normalize_distributor(row: dict[str, str], index: int, source_file: str) -> dict[str, str]:
    output = blank_row(index, source_file)
    title = first(row, "Track Title", "Title")
    period_start, period_end = parse_month(first(row, "Sales Month", "Period"))
    rights_type = first(row, "Rights Type").lower() or "master"
    income_type = first(row, "Income Type").lower() or "streaming"
    output.update(
        {
            "provider": "Distributor",
            "rights_type": rights_type,
            "income_type": income_type,
            "catalog_asset_id": asset_id("isrc", first(row, "ISRC"), title),
            "period_start": first(row, "Period Start") or period_start,
            "period_end": first(row, "Period End") or period_end,
            "payment_date": first(row, "Report Date", "Payment Date"),
            "territory": first(row, "Territory", "Country"),
            "platform_or_licensee": first(row, "Store", "DSP", "Platform", "Source"),
            "gross_amount": clean_amount(first(row, "Gross", "Gross Amount")),
            "deductions": clean_amount(first(row, "Fee", "Deductions")),
            "owner_net_amount": clean_amount(first(row, "Net", "Net Amount", "Net Payable")),
            "currency": first(row, "Currency") or "USD",
            "fx_rate": first(row, "FX", "FX Rate"),
            "notes": first(row, "Notes"),
        }
    )
    return output


def normalize_common_statement(
    row: dict[str, str],
    index: int,
    source_file: str,
    provider: str,
    rights_type: str,
    income_type: str,
) -> dict[str, str]:
    output = normalize_distributor(row, index, source_file)
    output.update(
        {
            "provider": provider,
            "rights_type": rights_type,
            "income_type": income_type,
        }
    )
    return output


def normalize_publisher_admin(row: dict[str, str], index: int, source_file: str) -> dict[str, str]:
    return normalize_common_statement(row, index, source_file, "Publisher Admin", "publishing", "mechanical")


def normalize_soundexchange(row: dict[str, str], index: int, source_file: str) -> dict[str, str]:
    return normalize_common_statement(row, index, source_file, "SoundExchange", "neighboring", "radio")


def normalize_direct_sync(row: dict[str, str], index: int, source_file: str) -> dict[str, str]:
    return normalize_common_statement(row, index, source_file, "Direct Sync", "sync", "sync")


def normalize_youtube(row: dict[str, str], index: int, source_file: str) -> dict[str, str]:
    output = blank_row(index, source_file)
    title = first(row, "Asset Title", "Track Title", "Title")
    report_date = parse_yyyymmdd(first(row, "date"))
    period_start, period_end = parse_month(first(row, "Reporting Month", "Month"))
    output.update(
        {
            "provider": "YouTube Content ID",
            "rights_type": "master",
            "income_type": "ugc",
            "catalog_asset_id": first(row, "custom_id") or asset_id("isrc", first(row, "ISRC", "isrc"), title),
            "period_start": period_start or report_date,
            "period_end": period_end or report_date,
            "territory": first(row, "Country", "Territory", "country_code"),
            "platform_or_licensee": "YouTube",
            "gross_amount": clean_amount(first(row, "Gross Revenue", "Gross", "youtube_revenue_split")),
            "owner_net_amount": clean_amount(first(row, "Net Revenue", "Net", "partner_revenue_usd")),
            "currency": first(row, "Currency", "currency_code") or "USD",
            "notes": f"content_type={first(row, 'Content Type')}" if first(row, "Content Type") else f"asset_id={first(row, 'asset_id')}" if first(row, "asset_id") else "",
        }
    )
    return output


def normalize_mlc(row: dict[str, str], index: int, source_file: str) -> dict[str, str]:
    output = blank_row(index, source_file)
    title = first(row, "Song Title", "Work Title", "Title")
    period_start, period_end = parse_month(first(row, "Usage Period", "Period"))
    output.update(
        {
            "provider": "The MLC",
            "rights_type": "publishing",
            "income_type": "mechanical",
            "catalog_asset_id": asset_id("iswc", first(row, "ISWC"), title),
            "period_start": first(row, "Period Start") or period_start,
            "period_end": first(row, "Period End") or period_end,
            "payment_date": first(row, "Report Date", "Payment Date"),
            "territory": first(row, "Territory") or "US",
            "platform_or_licensee": first(row, "DSP", "Platform", "Source"),
            "gross_amount": clean_amount(first(row, "Mechanical Gross", "Gross")),
            "deductions": clean_amount(first(row, "Admin Fee", "Deductions")),
            "owner_net_amount": clean_amount(first(row, "Publisher Net", "Net", "Net Payable")),
            "currency": first(row, "Currency") or "USD",
            "fx_rate": first(row, "FX", "FX Rate"),
            "notes": first(row, "Notes"),
        }
    )
    return output


def normalize_curve(row: dict[str, str], index: int, source_file: str) -> dict[str, str]:
    output = blank_row(index, source_file)
    title = first(row, "Track Title", "Work Title", "Title")
    rights_type = first(row, "Rights Type").lower() or "unknown"
    output.update(
        {
            "provider": "Curve",
            "rights_type": rights_type,
            "income_type": first(row, "Income Type").lower() or "unknown",
            "catalog_asset_id": first(row, "Source Work ID") or asset_id("isrc", first(row, "ISRC"), title),
            "period_start": first(row, "Period Start"),
            "period_end": first(row, "Period End"),
            "territory": first(row, "Territory"),
            "platform_or_licensee": first(row, "Royalty Source", "Source"),
            "gross_amount": clean_amount(first(row, "Gross")),
            "deductions": clean_amount(first(row, "Deductions")),
            "owner_net_amount": clean_amount(first(row, "Net Payable", "Net")),
            "currency": first(row, "Currency") or "USD",
        }
    )
    return output


def _filename_year_period(source_file: str) -> tuple[str, str]:
    match = re.search(r"(?<!\d)(20\d{2}|19\d{2})(?!\d)", source_file)
    if not match:
        return "", ""
    year = match.group(1)
    return f"{year}-01-01", f"{year}-12-31"


def normalize_bmi_annual(row: dict[str, str], index: int, source_file: str) -> dict[str, str]:
    """BMI annual consolidated XLSX. Sums Domestic Perf + Domestic Mech +
    International + Other into owner_net_amount and preserves the breakdown
    in `notes`. period_start/end is derived from the source filename year.
    """
    output = blank_row(index, source_file)
    title = first(row, "Title", "Work Title")
    work_id = first(row, "BMI Work #", "BMI Work#", "Work ID", "BMI Work")
    components = {
        "domestic_perf": _to_float(first(row, "Domestic Perf", "Domestic Performance")),
        "domestic_mech": _to_float(first(row, "Domestic Mech", "Domestic Mechanical")),
        "international": _to_float(first(row, "International (all CMOs)", "International")),
        "other": _to_float(first(row, "Other Income", "Other")),
    }
    total_value = _to_float(first(row, "Total"))
    if total_value is None:
        total_value = sum(v for v in components.values() if v is not None)
    components_note = ", ".join(
        f"{key}={value:.2f}" for key, value in components.items() if value is not None
    )
    period_start, period_end = _filename_year_period(source_file)
    output.update(
        {
            "provider": "BMI",
            "rights_type": "publishing",
            "income_type": "performance",
            "catalog_asset_id": (f"bmi:{work_id}" if work_id else asset_id("title", "", title)),
            "period_start": period_start,
            "period_end": period_end,
            "platform_or_licensee": "BMI consolidated",
            "owner_net_amount": f"{total_value:.2f}" if total_value is not None else "",
            "currency": "USD",
            "notes": components_note + (f"; bmi_work_id={work_id}" if work_id else ""),
        }
    )
    return output


def normalize_gema(row: dict[str, str], index: int, source_file: str) -> dict[str, str]:
    """GEMA German sub-publisher XLSX. Header layout typically:
    Title, ISWC, Distribution (EUR), Mechanical (EUR), Performance (EUR), Total (EUR).
    """
    output = blank_row(index, source_file)
    title = first(row, "Title", "Titel", "Work Title")
    iswc = first(row, "ISWC")
    total = clean_amount(first(row, "Total (EUR)", "Total", "Distribution (EUR)", "Distribution"))
    mechanical = clean_amount(first(row, "Mechanical (EUR)", "Mechanical"))
    performance = clean_amount(first(row, "Performance (EUR)", "Performance"))
    note_parts: list[str] = []
    if mechanical:
        note_parts.append(f"mechanical={mechanical}")
    if performance:
        note_parts.append(f"performance={performance}")
    period_start, period_end = _filename_year_period(source_file)
    output.update(
        {
            "provider": "GEMA",
            "rights_type": "publishing",
            "income_type": "performance",
            "catalog_asset_id": asset_id("iswc", iswc, title),
            "period_start": period_start,
            "period_end": period_end,
            "territory": "DE",
            "platform_or_licensee": "GEMA",
            "owner_net_amount": total,
            "currency": "EUR",
            "notes": "; ".join(note_parts),
        }
    )
    return output


def _to_float(value: str | None) -> float | None:
    if value in (None, ""):
        return None
    cleaned = str(value).replace("$", "").replace(",", "").strip()
    if not cleaned:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


PROFILES: dict[str, ProviderProfile] = {
    "ascap": ProviderProfile("ASCAP", "publishing", "performance", normalize_ascap),
    "bmi": ProviderProfile("BMI", "publishing", "performance", normalize_bmi),
    "bmi-annual": ProviderProfile("BMI Annual", "publishing", "performance", normalize_bmi_annual),
    "gema": ProviderProfile("GEMA", "publishing", "performance", normalize_gema),
    "distributor": ProviderProfile("Distributor", "master", "streaming", normalize_distributor),
    "publisher-admin": ProviderProfile("Publisher Admin", "publishing", "mechanical", normalize_publisher_admin),
    "soundexchange": ProviderProfile("SoundExchange", "neighboring", "radio", normalize_soundexchange),
    "direct-sync": ProviderProfile("Direct Sync", "sync", "sync", normalize_direct_sync),
    "youtube-content-id": ProviderProfile("YouTube Content ID", "master", "ugc", normalize_youtube),
    "mlc": ProviderProfile("The MLC", "publishing", "mechanical", normalize_mlc),
    "curve": ProviderProfile("Curve", "unknown", "unknown", normalize_curve),
}


def detect_delimiter(text_sample: str, requested: str) -> str:
    if requested != "auto":
        return requested
    try:
        dialect = csv.Sniffer().sniff(text_sample, delimiters=",\t|;")
        return dialect.delimiter
    except csv.Error:
        return ","


def read_rows(input_path: Path, delimiter: str, sheet: str | None = None) -> tuple[list[str], list[dict[str, str]]]:
    suffix = input_path.suffix.lower()
    if suffix in {".xlsx", ".xls", ".xlsm"}:
        return _read_xlsx_rows(input_path, sheet)
    raw = input_path.read_text(encoding="utf-8-sig")
    sample = raw[:8192] if raw else ""
    actual_delim = detect_delimiter(sample, delimiter)
    reader = csv.DictReader(io.StringIO(raw), delimiter=actual_delim)
    rows = [row for row in reader if has_meaningful_value(row)]
    headers = list(reader.fieldnames or [])
    return headers, rows


def _read_xlsx_rows(input_path: Path, sheet: str | None) -> tuple[list[str], list[dict[str, str]]]:
    try:
        import openpyxl  # type: ignore[import-not-found]
    except ImportError as exc:  # pragma: no cover - import-time guard
        raise SystemExit(
            "openpyxl is required for XLSX inputs. Install via: pip3 install -r requirements.txt"
        ) from exc
    workbook = openpyxl.load_workbook(input_path, data_only=True, read_only=True)
    if sheet and sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
    else:
        worksheet = workbook.active
    raw_rows = [row for row in worksheet.iter_rows(values_only=True)]
    header_idx = _detect_header_row(raw_rows)
    if header_idx is None:
        return [], []
    headers = [_to_str_cell(c) for c in raw_rows[header_idx]]
    headers = [h for h in headers if h]
    body: list[dict[str, str]] = []
    for raw in raw_rows[header_idx + 1:]:
        if not raw or not any(c is not None for c in raw):
            continue
        row_dict: dict[str, str] = {}
        for idx, header in enumerate(raw_rows[header_idx]):
            cell = raw[idx] if idx < len(raw) else None
            key = _to_str_cell(header)
            if not key:
                continue
            row_dict[key] = _to_str_cell(cell)
        if not has_meaningful_value(row_dict):
            continue
        body.append(row_dict)
    return headers, body


def _to_str_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _detect_header_row(rows: list[tuple]) -> int | None:
    """Find the first row that looks like column headers: ≥4 non-null string
    cells with no leading currency symbol. This skips preface metadata rows
    (e.g. "BMI Royalty Distribution — Consolidated Annual View 2022")."""
    for idx, row in enumerate(rows):
        if not row:
            continue
        non_null = [c for c in row if c is not None and str(c).strip()]
        if len(non_null) < 3:
            continue
        if not all(isinstance(c, str) for c in non_null):
            continue
        if any(str(c).strip().startswith(("$", "€", "£", "¥")) for c in non_null):
            continue
        # Title-rows usually contain only one cell + Nones; require fill on multiple columns.
        return idx
    return None


def load_column_map(path: Path | None) -> dict[str, str]:
    if path is None:
        return {}
    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, dict):
        raise SystemExit(f"--column-map must be a JSON object: {path}")
    valid: dict[str, str] = {}
    for source, canonical in data.items():
        if not isinstance(source, str) or not isinstance(canonical, str):
            continue
        if canonical not in CANONICAL_COLUMNS:
            raise SystemExit(
                f"--column-map target '{canonical}' is not a canonical ledger column. "
                f"Valid targets: {', '.join(sorted(CANONICAL_COLUMNS))}"
            )
        valid[source] = canonical
    return valid


def apply_column_map(input_row: dict[str, str], output_row: dict[str, str], column_map: dict[str, str]) -> None:
    """Fill any canonical fields that the profile left empty using the mapping."""
    for source_col, canonical_field in column_map.items():
        if output_row.get(canonical_field):
            continue
        source_value = first(input_row, source_col)
        if not source_value:
            continue
        if canonical_field in NUMERIC_FIELDS:
            output_row[canonical_field] = clean_amount(source_value)
        else:
            output_row[canonical_field] = source_value


def compute_population_rate(rows: list[dict[str, str]]) -> float:
    if not rows:
        return 0.0
    total = len(rows) * len(CRITICAL_FIELDS)
    filled = sum(
        1
        for row in rows
        for field in CRITICAL_FIELDS
        if (row.get(field) or "").strip()
    )
    return filled / total if total else 0.0


def expected_columns_not_found(provider: str, headers: list[str]) -> list[str]:
    aliases = FINANCIAL_ALIASES.get(provider, {})
    expected: set[str] = set()
    for field in CRITICAL_FIELDS + ("period_end", "platform_or_licensee"):
        for alias in aliases.get(field, []):
            expected.add(alias)
    if not expected:
        return []
    headers_lower = {h.lower() for h in headers if isinstance(h, str)}
    missing = [name for name in sorted(expected) if name.lower() not in headers_lower]
    return missing


def normalize(
    provider: str,
    input_path: Path,
    output_path: Path,
    delimiter: str = "auto",
    column_map_path: Path | None = None,
    sheet: str | None = None,
) -> dict[str, object]:
    profile = PROFILES.get(provider)
    if profile is None:
        supported = ", ".join(sorted(PROFILES))
        raise SystemExit(f"Unsupported provider '{provider}'. Supported providers: {supported}")

    column_map = load_column_map(column_map_path)
    headers, rows = read_rows(input_path, delimiter, sheet=sheet)

    normalized_rows: list[dict[str, str]] = []
    for index, row in enumerate(rows):
        output = profile.normalize(row, index + 1, str(input_path))
        if column_map:
            apply_column_map(row, output, column_map)
        normalized_rows.append(output)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=CANONICAL_COLUMNS)
        writer.writeheader()
        writer.writerows(normalized_rows)

    rate = compute_population_rate(normalized_rows)
    warnings: list[dict[str, object]] = []
    status = "ok"
    if normalized_rows and rate < PARTIAL_THRESHOLD:
        status = "partial"
        missing = expected_columns_not_found(provider, headers)
        warnings.append(
            {
                "profile": provider,
                "field_population_rate": round(rate, 4),
                "headers_seen": headers,
                "expected_columns_not_found": missing,
                "critical_fields_checked": list(CRITICAL_FIELDS),
                "hint": (
                    "Coverage of gross_amount/owner_net_amount/period_start fell below "
                    f"{int(PARTIAL_THRESHOLD * 100)}%. Pass --column-map mapping.json with "
                    "{source_column: canonical_field} or extend the provider profile."
                ),
            }
        )

    payload: dict[str, object] = {
        "status": status,
        "provider": profile.provider_name,
        "input": str(input_path),
        "output": str(output_path),
        "rows": len(normalized_rows),
        "field_population_rate": round(rate, 4),
        "headers_seen": headers,
        "warnings": warnings,
    }
    return payload


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--provider", required=True, help=f"One of: {', '.join(sorted(PROFILES))}")
    parser.add_argument("--input", required=True, help="Provider CSV/TSV export")
    parser.add_argument("--output", required=True, help="Output royalty-ledger.csv path")
    parser.add_argument(
        "--delimiter",
        default="auto",
        help="Field delimiter for the input file. Use 'auto' to sniff (default), or pass an explicit one ('\\t' for TSV).",
    )
    parser.add_argument(
        "--column-map",
        default=None,
        help="Optional JSON file mapping {source_column: canonical_field} for provider variants.",
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="Worksheet name for XLSX inputs (defaults to active sheet).",
    )
    args = parser.parse_args()

    delimiter = args.delimiter
    if delimiter != "auto":
        delimiter = bytes(delimiter, "utf-8").decode("unicode_escape")
    column_map_path = Path(args.column_map) if args.column_map else None

    payload = normalize(
        args.provider,
        Path(args.input),
        Path(args.output),
        delimiter=delimiter,
        column_map_path=column_map_path,
        sheet=args.sheet,
    )
    print(json.dumps(payload, indent=2))
    if payload["status"] != "ok":
        print(
            f"normalize-royalty-statement: status={payload['status']} — see warnings above.",
            file=sys.stderr,
        )
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
