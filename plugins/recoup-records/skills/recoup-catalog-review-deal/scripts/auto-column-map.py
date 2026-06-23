#!/usr/bin/env python3
"""Auto-generate a column map for normalize-royalty-statement.py.

When a seller's CSV/XLSX export does not match a provider profile's
expected column names, ``normalize-royalty-statement.py`` returns
``status: "partial"`` and tells the caller to pass
``--column-map mapping.json``. Customers will not know what that means.

This script writes that mapping automatically by fuzzy-matching the
input file's headers against:

1. The provider profile's known aliases (FINANCIAL_ALIASES from the
   normalizer — high-confidence semantic match).
2. The canonical ledger column names (lower-confidence catch-all).
3. A small dictionary of common synonyms ("net royalty", "amount paid",
   "publisher amount", etc.) that show up in real seller exports.

Output: a JSON object suitable for ``--column-map``, plus per-mapping
confidence scores. Exits non-zero only when confidence on the three
financially critical fields (gross_amount, owner_net_amount,
period_start) is below the configured threshold — at which point
human review actually is required.

Reads the input file's headers without depending on the normalizer's
internals: stdlib csv for CSV/TSV, openpyxl for XLSX (already an
optional dep documented in requirements.txt).
"""

from __future__ import annotations

import argparse
import csv
import io
import json
import os
import re
import sys
from difflib import SequenceMatcher
from pathlib import Path
from typing import Iterable

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# Re-use the normalizer's canonical schema and provider aliases so the two
# scripts cannot drift. Importing requires renaming the file with hyphens,
# so we exec it the same way the test harness does.
import importlib.util


def _load_normalizer() -> tuple[list[str], dict[str, dict[str, list[str]]]]:
    module_name = "normalize_royalty_statement"
    spec = importlib.util.spec_from_file_location(
        module_name,
        Path(__file__).resolve().parent / "normalize-royalty-statement.py",
    )
    if spec is None or spec.loader is None:
        raise SystemExit("Could not locate normalize-royalty-statement.py next to auto-column-map.py")
    module = importlib.util.module_from_spec(spec)
    # Register in sys.modules BEFORE exec_module so @dataclass(frozen=True)
    # can resolve cls.__module__ via sys.modules.get(...).
    sys.modules[module_name] = module
    spec.loader.exec_module(module)
    return module.CANONICAL_COLUMNS, module.FINANCIAL_ALIASES


CANONICAL_COLUMNS, FINANCIAL_ALIASES = _load_normalizer()


# Critical fields used as the gate. Mirrors the normalizer.
CRITICAL_FIELDS = ("gross_amount", "owner_net_amount", "period_start")
DEFAULT_CONFIDENCE_THRESHOLD = 0.65


# Common header synonyms across providers. The normalizer's per-provider
# aliases catch most cases; this covers seller-specific variants.
GENERAL_SYNONYMS: dict[str, list[str]] = {
    "gross_amount": [
        "gross",
        "gross amount",
        "gross revenue",
        "gross royalty",
        "amount",
        "amount paid",
        "royalty amount",
        "total amount",
        "earnings",
    ],
    "owner_net_amount": [
        "net",
        "net amount",
        "net payable",
        "net royalty",
        "net revenue",
        "publisher share",
        "publisher net",
        "publisher amount",
        "owner net",
        "owner share",
        "share to owner",
        "label net",
        "label share",
        "your share",
        "due to writer",
        "due to publisher",
    ],
    "gross_amount_inferred_from_net": [],  # placeholder; not a canonical column
    "period_start": [
        "period start",
        "start date",
        "period from",
        "from date",
        "earnings period start",
        "service period start",
        "sales month",
        "statement month",
        "usage period",
        "performance start",
        "reporting month",
        "month",
        "period",
    ],
    "period_end": [
        "period end",
        "end date",
        "period to",
        "to date",
        "service period end",
        "performance end",
        "reporting month",
    ],
    "payment_date": [
        "payment date",
        "paid date",
        "distribution date",
        "report date",
        "statement date",
    ],
    "territory": [
        "territory",
        "country",
        "country code",
        "region",
        "market",
    ],
    "platform_or_licensee": [
        "platform",
        "store",
        "dsp",
        "source",
        "licensee",
        "station",
        "network",
        "vendor",
        "service",
    ],
    "currency": [
        "currency",
        "currency code",
        "ccy",
    ],
    "fx_rate": [
        "fx",
        "fx rate",
        "exchange rate",
    ],
    "deductions": [
        "deductions",
        "fee",
        "fees",
        "admin fee",
        "admin",
        "commission",
    ],
    "participant_share": [
        "share %",
        "share pct",
        "split",
        "split %",
        "percentage",
    ],
    "pro_use_type": [
        "use type",
        "performance type",
        "perf type",
    ],
    "pro_credits": [
        "credits",
        "total credits",
        "premium credits",
    ],
    "pro_bonus_type": [
        "bonus type",
        "bonus category",
        "premium type",
    ],
    "cue_sheet_ref": [
        "cue sheet",
        "program",
    ],
}


def _normalize(text: str) -> str:
    """Lowercase and strip non-alphanumerics. Used for fuzzy matching."""
    return re.sub(r"[^a-z0-9]+", "", text.lower())


def _ratio(a: str, b: str) -> float:
    return SequenceMatcher(None, _normalize(a), _normalize(b)).ratio()


def best_match(
    header: str,
    candidates: Iterable[tuple[str, str]],
) -> tuple[str | None, float, str | None]:
    """Return (canonical_field, score, matched_alias) for the strongest
    candidate above the candidate set, or (None, 0, None) if nothing fires.

    Each candidate is a (canonical_field, alias_or_field) pair so we can
    record exactly which alias produced the best match.
    """
    best_field: str | None = None
    best_alias: str | None = None
    best_score = 0.0
    for field, alias in candidates:
        score = _ratio(header, alias)
        if score > best_score:
            best_field = field
            best_alias = alias
            best_score = score
    return best_field, best_score, best_alias


def gather_candidates(provider: str) -> list[tuple[str, str]]:
    """Build the (canonical_field, alias) candidate set for a provider.

    Order matters: provider-specific aliases come first so that — at the
    same fuzzy score — they win over generic synonyms.
    """
    candidates: list[tuple[str, str]] = []
    aliases = FINANCIAL_ALIASES.get(provider, {})
    for field, names in aliases.items():
        for name in names:
            candidates.append((field, name))
    for field, names in GENERAL_SYNONYMS.items():
        if field not in CANONICAL_COLUMNS:
            continue
        for name in names:
            candidates.append((field, name))
    # Direct canonical name matches as a final fallback.
    for field in CANONICAL_COLUMNS:
        candidates.append((field, field.replace("_", " ")))
    return candidates


def read_headers(input_path: Path, sheet: str | None) -> list[str]:
    suffix = input_path.suffix.lower()
    if suffix in {".xlsx", ".xls", ".xlsm"}:
        return _read_xlsx_headers(input_path, sheet)
    raw = input_path.read_text(encoding="utf-8-sig")
    sample = raw[:8192]
    delimiter = _sniff_delimiter(sample)
    reader = csv.reader(io.StringIO(raw), delimiter=delimiter)
    for row in reader:
        if any((cell or "").strip() for cell in row):
            return [(cell or "").strip() for cell in row]
    return []


def _sniff_delimiter(sample: str) -> str:
    try:
        return csv.Sniffer().sniff(sample, delimiters=",\t|;").delimiter
    except csv.Error:
        return ","


def _read_xlsx_headers(input_path: Path, sheet: str | None) -> list[str]:
    try:
        import openpyxl  # type: ignore[import-not-found]
    except ImportError as exc:
        raise SystemExit(
            "openpyxl is required for XLSX inputs. Install: pip3 install openpyxl"
        ) from exc
    workbook = openpyxl.load_workbook(input_path, data_only=True, read_only=True)
    if sheet and sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
    else:
        worksheet = workbook.active
    for row in worksheet.iter_rows(values_only=True):
        if not row:
            continue
        non_null = [c for c in row if c is not None and str(c).strip()]
        if len(non_null) < 3:
            continue
        if not all(isinstance(c, str) for c in non_null):
            continue
        return [str(c).strip() if c is not None else "" for c in row]
    return []


def build_mapping(
    headers: list[str],
    provider: str,
    threshold: float,
) -> dict:
    candidates = gather_candidates(provider)
    suggestions: dict[str, dict] = {}
    used_canonical: set[str] = set()
    rejected: list[dict] = []

    # Score every header once, then claim canonicals greedily by score
    # so the same canonical is never used twice.
    scored: list[tuple[float, str, str | None, str | None]] = []
    for header in headers:
        if not header:
            continue
        field, score, alias = best_match(header, candidates)
        scored.append((score, header, field, alias))

    scored.sort(reverse=True)
    column_map: dict[str, str] = {}
    for score, header, field, alias in scored:
        if not field:
            continue
        if field in used_canonical:
            rejected.append(
                {
                    "header": header,
                    "would_have_mapped_to": field,
                    "score": round(score, 3),
                    "reason": "canonical_already_claimed",
                }
            )
            continue
        if score < threshold:
            rejected.append(
                {
                    "header": header,
                    "would_have_mapped_to": field,
                    "score": round(score, 3),
                    "reason": "below_threshold",
                }
            )
            continue
        column_map[header] = field
        suggestions[header] = {
            "canonical": field,
            "score": round(score, 3),
            "matched_alias": alias,
        }
        used_canonical.add(field)

    critical_status: dict[str, dict] = {}
    for field in CRITICAL_FIELDS:
        mapped_header = next(
            (h for h, info in suggestions.items() if info["canonical"] == field),
            None,
        )
        if mapped_header:
            critical_status[field] = {
                "status": "ok",
                "header": mapped_header,
                "score": suggestions[mapped_header]["score"],
            }
        else:
            critical_status[field] = {"status": "missing"}

    confident = all(info["status"] == "ok" for info in critical_status.values())

    return {
        "status": "ok" if confident else "needs_review",
        "provider": provider,
        "threshold": threshold,
        "headers_seen": headers,
        "column_map": column_map,
        "suggestions": suggestions,
        "rejected": rejected,
        "critical_fields": critical_status,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--provider", required=True, help="Provider profile key (ascap, bmi, mlc, distributor, ...).")
    parser.add_argument("--input", required=True, help="Provider CSV/TSV/XLSX export.")
    parser.add_argument(
        "--output",
        default=None,
        help="Where to write the column-map JSON. Default: print to stdout only.",
    )
    parser.add_argument(
        "--threshold",
        type=float,
        default=DEFAULT_CONFIDENCE_THRESHOLD,
        help=f"Minimum fuzzy match score (default {DEFAULT_CONFIDENCE_THRESHOLD}).",
    )
    parser.add_argument("--sheet", default=None, help="Worksheet name for XLSX inputs.")
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.is_file():
        raise SystemExit(f"Input file not found: {input_path}")

    headers = read_headers(input_path, args.sheet)
    if not headers:
        payload = {
            "status": "no_headers",
            "input": str(input_path),
            "headers_seen": [],
            "column_map": {},
        }
        print(json.dumps(payload, indent=2))
        return 1

    payload = build_mapping(headers, args.provider, args.threshold)
    payload["input"] = str(input_path)

    if args.output:
        out_path = Path(args.output)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        # Write the bare {source -> canonical} object; this is what the
        # normalizer expects from --column-map. The full diagnostic
        # payload (suggestions, rejected, etc.) goes to stdout.
        out_path.write_text(json.dumps(payload["column_map"], indent=2) + "\n", encoding="utf-8")
        payload["output"] = str(out_path)

    print(json.dumps(payload, indent=2))
    return 0 if payload["status"] == "ok" else 1


if __name__ == "__main__":
    raise SystemExit(main())
